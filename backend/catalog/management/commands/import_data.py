"""
Management command: import_data

Usage:
    python manage.py import_data <path-to-xlsx> [--flush] [--dry-run]

Reads four sheets from the workbook in order:
    Notes → Note_Aliases → Perfumes → Perfume-Notes

Sheet columns:
    Notes         : Note_ID, Note_Name
    Note_Aliases  : Alias_Name, Note_ID, Canonical_Note_Name, Source
    Perfumes      : ID, Perfume_Name, Brand, Concentration, Main_Accords, URL
    Perfume-Notes : Perfume_ID, Note_ID, Layer

Notes:
    - Accords are NOT imported as a normalized table (no Accords / Accord-Note
      sheets in this workbook); the Accord/AccordNote models stay empty.
    - ``Main_Accords`` IS stored on ``Perfume.main_accords`` (denormalized display
      string) to power accord chips on catalog cards.

Normalises sentinels:
    - Text == "Undefined" / blank → NULL / blank.

--flush:
    Wipes the catalog AND user profiles before importing, in FK-safe order.
    Use this when the workbook ids were renumbered (the canonical reload path):
    upserting would otherwise leave stale rows and orphaned profile note links.

--dry-run:
    Reports row counts inside a transaction that is rolled back; nothing is written.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalog.models import Accord, AccordNote, Note, NoteAlias, Perfume, PerfumeNote

BATCH = 2000  # rows per bulk_create call


def _clean(value: object) -> str | None:
    """Return None for blank / Undefined sentinel; str otherwise."""
    if value is None:
        return None
    s = str(value).strip()
    return None if s in ("", "Undefined") else s


class Command(BaseCommand):
    help = "Import perfume data from the project Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("xlsx", type=Path, help="Path to Perfume_Database.xlsx")
        parser.add_argument(
            "--flush",
            action="store_true",
            help="Wipe catalog + profiles before importing (canonical reload).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Report row counts without writing to the database.",
        )

    def handle(self, *args, **options):
        xlsx: Path = options["xlsx"]
        flush: bool = options["flush"]
        dry_run: bool = options["dry_run"]

        if not xlsx.exists():
            raise CommandError(f"File not found: {xlsx}")

        self.stdout.write(f"Opening {xlsx} …")
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

        try:
            stats = self._import(wb, flush, dry_run)
        finally:
            wb.close()

        self.stdout.write("\n" + "=" * 60)
        self.stdout.write("SUMMARY" + (" (dry-run — nothing written)" if dry_run else ""))
        self.stdout.write("=" * 60)
        for table, counts in stats.items():
            self.stdout.write(
                f"  {table:20s}  created={counts['created']:>6}  "
                f"updated={counts['updated']:>6}  skipped={counts['skipped']:>6}"
            )

    # ------------------------------------------------------------------
    def _import(self, wb: openpyxl.Workbook, flush: bool, dry_run: bool) -> dict:
        stats: dict[str, dict] = {}

        with transaction.atomic():
            if flush:
                self._flush()

            stats["Note"] = self._load_notes(wb, dry_run)
            stats["NoteAlias"] = self._load_aliases(wb, dry_run)
            stats["Perfume"] = self._load_perfumes(wb, dry_run)
            stats["PerfumeNote"] = self._load_perfume_notes(wb, dry_run)

            if dry_run:
                transaction.set_rollback(True)

        return stats

    # ------------------------------------------------------------------
    def _flush(self) -> None:
        """Erase catalog + profiles in FK-safe order before a clean reload."""
        from profiles.models import Profile, ProfileGroup, ProfileNote

        self.stdout.write("Flushing existing data …")
        for model in (
            ProfileNote,
            Profile,
            ProfileGroup,
            PerfumeNote,
            NoteAlias,
            AccordNote,
            Perfume,
            Note,
            Accord,
        ):
            deleted, _ = model.objects.all().delete()
            self.stdout.write(f"  flushed {model.__name__:14s} ({deleted} rows)")

    # ------------------------------------------------------------------
    def _load_notes(self, wb, dry_run: bool) -> dict:
        ws = wb["Notes"]
        created = updated = skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            note_id, note_name = row[0], row[1]
            if not note_id:
                skipped += 1
                continue
            _, was_created = Note.objects.update_or_create(
                note_id=str(note_id).strip(),
                defaults={"name": str(note_name).strip()},
            )
            created += int(was_created)
            updated += int(not was_created)

        self.stdout.write(f"  Notes:        created={created}  updated={updated}  skipped={skipped}")
        return {"created": created, "updated": updated, "skipped": skipped}

    # ------------------------------------------------------------------
    def _load_aliases(self, wb, dry_run: bool) -> dict:
        if "Note_Aliases" not in wb.sheetnames:
            self.stdout.write("  Note_Aliases: sheet not present — skipped")
            return {"created": 0, "updated": 0, "skipped": 0}

        ws = wb["Note_Aliases"]
        created = skipped = 0

        note_map = {n.note_id: n for n in Note.objects.only("id", "note_id")}
        canonical_names = {
            name.casefold()
            for name in Note.objects.values_list("name", flat=True)
        }
        seen_alias: set[str] = set()
        batch: list[NoteAlias] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            alias_name = _clean(row[0])
            note_id = str(row[1]).strip() if row[1] else None
            source = _clean(row[3]) if len(row) > 3 else None

            if not alias_name or not note_id:
                skipped += 1
                continue

            note = note_map.get(note_id)
            if note is None:
                skipped += 1
                self.stderr.write(f"  WARN NoteAlias: unknown note_id={note_id} (alias={alias_name})")
                continue

            key = alias_name.casefold()
            if key in canonical_names:
                skipped += 1
                self.stderr.write(f"  WARN NoteAlias: alias collides with canonical name: {alias_name}")
                continue
            if key in seen_alias:
                skipped += 1
                continue
            seen_alias.add(key)

            batch.append(NoteAlias(alias_name=alias_name, note=note, source=source))

            if len(batch) >= BATCH:
                created += self._flush_aliases(batch)
                batch.clear()

        if batch:
            created += self._flush_aliases(batch)

        self.stdout.write(f"  NoteAliases:  created={created}  updated={0}  skipped={skipped}")
        return {"created": created, "updated": 0, "skipped": skipped}

    def _flush_aliases(self, batch: list[NoteAlias]) -> int:
        result = NoteAlias.objects.bulk_create(batch, ignore_conflicts=True)
        return len(result)

    # ------------------------------------------------------------------
    def _load_perfumes(self, wb, dry_run: bool) -> dict:
        ws = wb["Perfumes"]
        created = updated = skipped = 0
        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                skipped += 1
                continue
            perfume_id = str(row[0]).strip()
            name = _clean(row[1]) or "(unnamed)"
            brand = _clean(row[2])
            concentration = _clean(row[3])
            main_accords = _clean(row[4])  # denormalized display string (accord chips)
            url = _clean(row[5]) if len(row) > 5 else None

            _, was_created = Perfume.objects.update_or_create(
                perfume_id=perfume_id,
                defaults={
                    "name": name,
                    "brand": brand,
                    "concentration": concentration,
                    "main_accords": main_accords,
                    "url": url,
                },
            )
            created += int(was_created)
            updated += int(not was_created)

            total += 1
            if total % 5000 == 0:
                self.stdout.write(f"    … {total} perfumes processed")

        self.stdout.write(f"  Perfumes:     created={created}  updated={updated}  skipped={skipped}")
        return {"created": created, "updated": updated, "skipped": skipped}

    # ------------------------------------------------------------------
    def _load_perfume_notes(self, wb, dry_run: bool) -> dict:
        ws = wb["Perfume-Notes"]
        created = skipped = 0

        perfume_map = {p.perfume_id: p for p in Perfume.objects.only("id", "perfume_id")}
        note_map = {n.note_id: n for n in Note.objects.only("id", "note_id")}

        valid_layers = {"top", "middle", "base"}
        batch: list[PerfumeNote] = []
        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            perfume_id, note_id, layer = row[0], row[1], row[2]
            perfume_id = str(perfume_id).strip() if perfume_id else None
            note_id = str(note_id).strip() if note_id else None
            layer = str(layer).strip().lower() if layer else None

            if layer not in valid_layers:
                skipped += 1
                continue

            perfume = perfume_map.get(perfume_id)
            note = note_map.get(note_id)

            if not perfume or not note:
                skipped += 1
                if not perfume:
                    self.stderr.write(f"  WARN PerfumeNote: unknown perfume_id={perfume_id}")
                if not note:
                    self.stderr.write(f"  WARN PerfumeNote: unknown note_id={note_id}")
                continue

            batch.append(PerfumeNote(perfume=perfume, note=note, layer=layer))

            if len(batch) >= BATCH:
                created += self._flush_perfume_notes(batch)
                batch.clear()

            total += 1
            if total % 50000 == 0:
                self.stdout.write(f"    … {total} perfume-notes processed")

        if batch:
            created += self._flush_perfume_notes(batch)

        self.stdout.write(f"  PerfumeNote:  created={created}  updated={0}  skipped={skipped}")
        return {"created": created, "updated": 0, "skipped": skipped}

    def _flush_perfume_notes(self, batch: list[PerfumeNote]) -> int:
        result = PerfumeNote.objects.bulk_create(batch, ignore_conflicts=True)
        return len(result)
